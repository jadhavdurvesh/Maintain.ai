"""Generates a multi-sheet Excel workbook covering machines, reliability,
failure analysis, work orders, spare parts, and active alerts."""
import io
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from .. import models

HEADER_FILL = PatternFill(start_color="1F2733", end_color="1F2733", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(wb, title, headers, rows):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    for col_idx, header in enumerate(headers, start=1):
        width = max(12, min(40, len(str(header)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    return ws


def build_excel_report(db: Session) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    machines = db.query(models.Machine).all()

    _write_sheet(
        wb, "Machines",
        ["Code", "Name", "Category", "Location", "Department", "Operating Hours", "Health Score", "Status", "Criticality"],
        [[m.machine_code, m.name, m.category, m.location, m.department, m.operating_hours, m.health_score, m.status, m.criticality] for m in machines],
    )

    rel_rows = []
    for m in machines:
        faults = db.query(models.FaultRecord).filter_by(machine_id=m.id).count()
        records = db.query(models.MaintenanceRecord).filter_by(machine_id=m.id).all()
        completed = sum(1 for r in records if r.status == models.MaintenanceStatus.completed)
        rate = round(completed / len(records) * 100, 1) if records else None
        rel_rows.append([m.name, faults, len(records), completed, rate, m.health_score])
    _write_sheet(
        wb, "Reliability",
        ["Machine", "Faults", "Maintenance Total", "Completed", "Completion %", "Health Score"],
        rel_rows,
    )

    faults = db.query(models.FaultRecord).all()
    cause_counts = Counter((f.cause or "unspecified") for f in faults).most_common(20)
    _write_sheet(wb, "Failure Analysis", ["Cause", "Occurrences"], [[c, n] for c, n in cause_counts])

    work_orders = db.query(models.WorkOrder).all()
    machine_names = {m.id: m.name for m in machines}
    _write_sheet(
        wb, "Work Orders",
        ["ID", "Machine", "Problem", "Priority", "Status", "Assigned To", "Created At"],
        [[wo.id, machine_names.get(wo.machine_id, wo.machine_id), wo.problem, wo.priority, wo.status,
          wo.assigned_to, wo.created_at.strftime("%Y-%m-%d %H:%M") if wo.created_at else ""] for wo in work_orders],
    )

    parts = db.query(models.SparePart).all()
    _write_sheet(
        wb, "Spare Parts",
        ["Name", "Part Number", "Quantity", "Minimum Stock", "Compatible Categories"],
        [[p.name, p.part_number, p.quantity, p.minimum_stock, p.compatible_machine_categories] for p in parts],
    )

    alerts = db.query(models.Alert).filter_by(resolved=False).all()
    _write_sheet(
        wb, "Active Alerts",
        ["Machine", "Type", "Severity", "Message", "Created At"],
        [[machine_names.get(a.machine_id, a.machine_id), a.alert_type, a.severity, a.message,
          a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else ""] for a in alerts],
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
